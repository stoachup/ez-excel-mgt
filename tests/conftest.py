import pytest
import openpyxl


@pytest.fixture(scope="function")
def create_test_excel(tmp_path):
    """Fixture to create a temporary Excel file for testing."""
    # Create a temporary file path for the Excel file
    excel_path = tmp_path / "test.xlsx"
    
    # Create a new workbook and add a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    
    # Prepopulate the sheet with some data
    sheet["A1"] = "First row"
    sheet["A2"] = "Second row"
    sheet["A3"] = "Name"
    sheet["B3"] = "Age"
    sheet["C3"] = "Gender"
    sheet["A4"] = "Irène"
    sheet["B4"] = 25
    sheet["A5"] = "Matthieu"
    sheet["B5"] = 26

    # Write to the temporary file
    workbook.save(excel_path)
    
    return str(excel_path), sheet.title, 3


@pytest.fixture(scope="function")
def create_test_excel_float(tmp_path):
    """Fixture to create a temporary Excel file for testing."""
    # Create a temporary file path for the Excel file
    excel_path = tmp_path / "test.xlsx"
    
    # Create a new workbook and add a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    
    # Prepopulate the sheet with some data
    sheet["A1"] = "First row"
    sheet["A2"] = "Second row"
    sheet["A3"] = "Name"
    sheet["B3"] = "Age"
    sheet["C3"] = "Gender"
    sheet["D3"] = "Revenue"
    sheet["A4"] = "Irène"
    sheet["B4"] = 25
    sheet["C4"] = "F"
    sheet["D4"] = 1234.56
    sheet["A5"] = "Matthieu"
    sheet["B5"] = 26
    sheet["C5"] = "M"
    sheet["D5"] = 9876.54

    # Write to the temporary file
    workbook.save(excel_path)
    
    return str(excel_path), sheet.title, 3


@pytest.fixture(scope="function")
def create_test_excel_commented(tmp_path):
    """Fixture to create a temporary Excel file for testing with comments after the header row."""
    # Create a temporary file path for the Excel file
    excel_path = tmp_path / "test.xlsx"
    
    # Create a new workbook and add a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    
    # Prepopulate the sheet with some data
    sheet["A1"] = "First row"
    sheet["A2"] = "Second row"
    sheet["A3"] = "Name"
    sheet["B3"] = "Age"
    sheet["C3"] = "Gender"
    sheet["A4"] = "Commented name"
    sheet["B4"] = "Commented age"
    sheet["C4"] = "Commented gender"

    # Write to the temporary file
    workbook.save(excel_path)
    
    return str(excel_path), sheet.title, 3


@pytest.fixture(scope="function")
def create_empty_test_excel(tmp_path):
    """Fixture to create a temporary EmptyExcel file for testing."""
    # Create a temporary file path for the Excel file
    excel_path = "dest.xlsx"
    
    # Create a new workbook and add a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Test Sheet"
    
    # Write to the temporary file
    workbook.save(excel_path)
    
    return str(excel_path), sheet.title, 0


@pytest.fixture(scope="function")
def create_test_excel_with_data_to_aggregate(tmp_path):
    """Fixture to create a temporary Excel file for testing."""
    # Create a temporary file path for the Excel file
    excel_path = tmp_path / "test.xlsx"
    
    # Create a new workbook and add a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    
    # Prepopulate the sheet with some data
    sheet["A1"] = "First"
    sheet["B1"] = "Second"
    sheet["C1"] = "Third"
    sheet["D1"] = "Fourth"
    sheet["A2"] = 1
    sheet["B2"] = 2
    sheet["C2"] = 3
    sheet["D2"] = 4
    sheet["A3"] = 5
    sheet["B3"] = 6
    sheet["C3"] = 7
    sheet["D3"] = 8

    # Write to the temporary file
    workbook.save(excel_path)
    
    return str(excel_path), sheet.title, 1


@pytest.fixture(scope="function")
def create_test_excel_with_3_sheets(tmp_path):
    """Fixture to create a temporary Excel file for testing with 3 sheets."""
    # Create a temporary file path for the Excel file
    excel_path = tmp_path / "test.xlsx"
    
    # Create a new workbook and add a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "New Sheet 1"
    workbook.create_sheet("New Sheet 2")
    workbook.create_sheet("New Sheet 3")
    
    # Write to the temporary file
    workbook.save(excel_path)
    
    return str(excel_path), ["New Sheet 1", "New Sheet 2", "New Sheet 3"], 3


