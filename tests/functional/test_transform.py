# This project uses Poetry for dependency management.
from pathlib import Path
import openpyxl
from ez_excel_mgt import ExcelTemplate


def test_transform_sum_row(create_test_excel_with_data_to_aggregate, create_empty_test_excel):
    """Test copy."""
    source_file_path, source_sheet_name, _header_row = create_test_excel_with_data_to_aggregate
    dest_file_path, dest_sheet_name, _header_row = create_empty_test_excel

    template = ExcelTemplate(dest_file_path)
    template.goto_sheet(dest_sheet_name, cell=(1, 1))
    template.aggregate_range_from(source_file_path, source_sheet_name, ((2, 1), (3, 4)), action='sum', mode='row')
    template.save(dest_file_path)

    assert Path(dest_file_path).exists()

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(dest_file_path)
    sheet = workbook[dest_sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert sheet["A1"].value == 10
    assert sheet["A2"].value == 26

def test_transform_sum_col(create_test_excel_with_data_to_aggregate, create_empty_test_excel):
    """Test copy."""
    source_file_path, source_sheet_name, _header_row = create_test_excel_with_data_to_aggregate
    dest_file_path, dest_sheet_name, _header_row = create_empty_test_excel
    
    template = ExcelTemplate(dest_file_path)
    template.goto_sheet(dest_sheet_name, cell=(1, 1))
    template.aggregate_range_from(source_file_path, source_sheet_name, ((2, 1), (3, 4)), action='sum', mode='col')
    template.save(dest_file_path)

    assert Path(dest_file_path).exists()

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(dest_file_path)
    sheet = workbook[dest_sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert sheet["A1"].value == 6
    assert sheet["B1"].value == 8
    assert sheet["C1"].value == 10
    assert sheet["D1"].value == 12

def test_transform_avg_row(create_test_excel_with_data_to_aggregate, create_empty_test_excel):
    """Test copy."""
    source_file_path, source_sheet_name, _header_row = create_test_excel_with_data_to_aggregate
    dest_file_path, dest_sheet_name, _header_row = create_empty_test_excel

    template = ExcelTemplate(dest_file_path)
    template.goto_sheet(dest_sheet_name, cell=(1, 1))
    template.aggregate_range_from(source_file_path, source_sheet_name, ((2, 1), (3, 4)), action='avg', mode='row')
    template.save(dest_file_path)

    assert Path(dest_file_path).exists()

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(dest_file_path)
    sheet = workbook[dest_sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert sheet["A1"].value == 2.5
    assert sheet["A2"].value == 6.5

def test_transform_avg_col(create_test_excel_with_data_to_aggregate, create_empty_test_excel):
    """Test copy."""
    source_file_path, source_sheet_name, _header_row = create_test_excel_with_data_to_aggregate
    dest_file_path, dest_sheet_name, _header_row = create_empty_test_excel


    template = ExcelTemplate(dest_file_path)
    template.goto_sheet(dest_sheet_name, cell=(1, 1))
    template.aggregate_range_from(source_file_path, source_sheet_name, ((2, 1), (3, 4)), action='avg', mode='col')
    template.save(dest_file_path)

    assert Path(dest_file_path).exists()

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(dest_file_path)
    sheet = workbook[dest_sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert sheet["A1"].value == 3
    assert sheet["B1"].value == 4
    assert sheet["C1"].value == 5
    assert sheet["D1"].value == 6
