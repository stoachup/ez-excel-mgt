# This project uses Poetry for dependency management.
import shutil
from pathlib import Path
import openpyxl
from ez_excel_mgt import ExcelTemplate

def test_transpose_excel_range(create_test_excel, create_empty_test_excel):
    """Test on demand input files."""
    source_file_path = '/Users/christophe/ENTSOE/PEMMDB-utils/demand/input/pemmdb_DE00_National-Trends_WO2024.xlsx'
    source_sheet_name = 'Annual Demand'
    
    dest_file_path = '/Users/christophe/ENTSOE/PEMMDB-utils/demand/output/test_demand.xlsx'
    dest_sheet_name = 'Annual Demand'

    template_file_path = '/Users/christophe/ENTSOE/PEMMDB-utils/demand/test.xlsx'
    shutil.copy(template_file_path, dest_file_path)

    template = ExcelTemplate(dest_file_path)
    template.goto_sheet(dest_sheet_name, cell='B4')
    template.copy_range_from(source_file_path, source_sheet_name, 'F6:AJ11', True, None)
    template.save(dest_file_path)

    assert Path(dest_file_path).exists()

    # Load the modified Excel file and verify the contents
    source_workbook = openpyxl.load_workbook(source_file_path)
    source_sheet = source_workbook[source_sheet_name]

    # Load the modified Excel file and verify the contents
    dest_workbook = openpyxl.load_workbook(dest_file_path)
    dest_sheet = dest_workbook[dest_sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert dest_sheet["B4"].value == source_sheet["F6"].value
    assert dest_sheet["C4"].value == source_sheet["F7"].value

def test_transpose_row_col(create_test_excel, create_empty_test_excel):
    """Test on demand input files."""
    source_file_path = '/Users/christophe/ENTSOE/PEMMDB-utils/demand/input/pemmdb_DE00_National-Trends_WO2024.xlsx'
    source_sheet_name = 'Annual Demand'
    
    dest_file_path = '/Users/christophe/ENTSOE/PEMMDB-utils/demand/output/test_demand.xlsx'
    dest_sheet_name = 'Annual Demand'

    template_file_path = '/Users/christophe/ENTSOE/PEMMDB-utils/demand/test.xlsx'
    shutil.copy(template_file_path, dest_file_path)

    template = ExcelTemplate(dest_file_path)
    template.goto_sheet(dest_sheet_name, cell=(4, 2))
    template.copy_range_from(source_file_path, source_sheet_name, ((6, 6), (11, 36)), True, None)
    template.save(dest_file_path)

    assert Path(dest_file_path).exists()

    # Load the modified Excel file and verify the contents
    source_workbook = openpyxl.load_workbook(source_file_path)
    source_sheet = source_workbook[source_sheet_name]

    # Load the modified Excel file and verify the contents
    dest_workbook = openpyxl.load_workbook(dest_file_path)
    dest_sheet = dest_workbook[dest_sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert dest_sheet["B4"].value == source_sheet["F6"].value
    assert dest_sheet["C4"].value == source_sheet["F7"].value
