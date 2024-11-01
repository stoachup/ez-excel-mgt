# This project uses Poetry for dependency management.
 
import polars as pl
import pandas as pd
import pytest
import openpyxl
from ez_excel_mgt import ExcelTemplate


def generate_test_data(data, data_type="polars"):
    """Utility to generate test data in different formats (Polars, Pandas, or dict of lists)."""
    
    if data_type == "polars":
        return pl.DataFrame(data)
    elif data_type == "pandas":
        return pd.DataFrame(data)
    elif data_type == "dict":
        return data
    elif data_type == "list":
        return list(data.values())
    else:
        raise ValueError(f"Unsupported data type: {data_type}")
    

@pytest.mark.parametrize("data_type", ["polars", "pandas", "dict", "list"])
def test_fill_sheet(create_test_excel, data_type):
    """Test inserting data at the end of the sheet."""
    excel_path, sheet_name, header_row = create_test_excel
    
    # Create a sample Polars DataFrame
    df = generate_test_data({
        "Name": ["Alice", "Bob"],
        "Age": [25, 30],
        "Gender": ["F", "M"],
    }, data_type)

    kwargs = {}
    if data_type == "list":
        kwargs["columns"] = ["Name", "Age", "Gender"]

    template = ExcelTemplate(excel_path)
    template.goto_sheet(sheet_name)
    template.set_header_location((header_row, 1), 'row')
    template.fill_with(df, **kwargs)
    template.save(excel_path)

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheet_name]
    
    # Assert that data is inserted with a header row (named columns)
    assert sheet["A3"].value == "Name"
    assert sheet["B3"].value == "Age"
    assert sheet["C3"].value == "Gender"
    assert sheet["A6"].value == "Alice"
    assert sheet["B6"].value == 25
    assert sheet["C6"].value == "F"
    assert sheet["A7"].value == "Bob"
    assert sheet["B7"].value == 30
    assert sheet["C7"].value == "M"

def test_fill_sheet_with_list_of_lists_with_no_columns(create_test_excel):
    """Test inserting data at the end of the sheet."""
    excel_path, sheet_name, header_row = create_test_excel
    
    # Create a sample Polars DataFrame
    df = generate_test_data({
        "Name": ["Alice", "Bob"],
        "Age": [25, 30],
        "Gender": ["F", "M"]
    }, 'list')

    kwargs = {}

    with pytest.raises(ValueError, match="Column names must be provided for List of Lists."):
        template = ExcelTemplate(excel_path)
        template.goto_sheet(sheet_name)
        template.set_header_location((header_row, 1), 'row')
        template.fill_with(df, **kwargs)


def test_fill_sheet_with_list_of_lists_with_not_enough_columns(create_test_excel):
    """Test inserting data at the end of the sheet."""
    excel_path, sheet_name, header_row = create_test_excel
    
    # Create a sample Polars DataFrame
    df = generate_test_data({
        "Name": ["Alice", "Bob"],
        "Age": [25, 30],
        "Gender": ["F", "M"]
    }, 'list')

    kwargs = {}
    kwargs["columns"] = ["Age", "Gender"]

    with pytest.raises(ValueError, match="List of columns and list of lists have different lengths."):
        template = ExcelTemplate(excel_path)
        template.goto_sheet(sheet_name)
        template.set_header_location((header_row, 1), 'row')
        template.fill_with(df, **kwargs)


@pytest.mark.parametrize("data_type", ["dict", "list"])
def test_fill_sheet_with_list_or_dict_of_lists_of_unequal_length(create_test_excel, data_type):
    """Test behavior when a column in the DataFrame is not of the same length as the others."""
    excel_path, sheet_name, header_row = create_test_excel
    
    # Create a DataFrame with mismatching column names
    df = generate_test_data({
        "Name": ["Alice", "Bob"],
        "Age": [25, 30, 54],
        "Gender": ["F", "M"]
    }, data_type)

    kwargs = {}
    if data_type == "list":
        kwargs["columns"] = ["Name", "Age", "Gender"]

    # Expect an error or handle the mismatch gracefully (depending on the implementation)
    message = "At least one list in the (dictionary|list) of lists has a different length than the others."
    with pytest.raises(ValueError, match=message):
        template = ExcelTemplate(excel_path)
        template.goto_sheet(sheet_name)
        template.set_header_location((header_row, 1), 'row')
        template.fill_with(df, **kwargs)


@pytest.mark.parametrize("data_type", ["pandas", "polars", "dict", "list"])
def test_fill_sheet_with_strict_and_missing_column(create_test_excel, data_type):
    """Test behavior when a column in the DataFrame is not of the same length as the others."""
    excel_path, sheet_name, header_row = create_test_excel
    
    # Create a DataFrame with mismatching column names
    df = generate_test_data({
        "Age": [25, 30],
        "Gender": ["F", "M"]
    }, data_type)

    kwargs = { "strict": True }
    if data_type == "list":
        kwargs["columns"] = ["Age", "Gender"]

    # Expect an error or handle the mismatch gracefully (depending on the implementation)
    message = "Header 'Name' in Sheet1 in the ExcelTemplate is missing in the DataFrame."
    with pytest.raises(ValueError, match=message):
        template = ExcelTemplate(excel_path)
        template.goto_sheet(sheet_name)
        template.set_header_location((header_row, 1), 'row')
        template.fill_with(df, **kwargs)


@pytest.mark.parametrize("data_type", ["pandas", "polars", "dict", "list"])
def test_fill_sheet_with_strict_and_unfound_column(create_test_excel, data_type):
    """Test behavior when a column in the DataFrame is not of the same length as the others."""
    excel_path, sheet_name, header_row = create_test_excel
    
    # Create a DataFrame with mismatching column names
    df = generate_test_data({
        "Full Name": ["Alice", "Bob"],
        "Age": [25, 30],
        "Gender": ["F", "M"]
    }, data_type)

    kwargs = { "strict": True }
    if data_type == "list":
        kwargs["columns"] = ["Full Name", "Age", "Gender"]

    # Expect an error or handle the mismatch gracefully (depending on the implementation)
    message = "Header 'Name' in Sheet1 in the ExcelTemplate is missing in the DataFrame."
    with pytest.raises(ValueError, match=message):
        template = ExcelTemplate(excel_path)
        template.goto_sheet(sheet_name)
        template.set_header_location((header_row, 1), 'row')
        template.fill_with(df, **kwargs)


@pytest.mark.parametrize("data_type", ["polars", "pandas", "dict", "list"])
def test_fill_sheet_with_overwrite_longer_than_template(create_test_excel, data_type):
    """Test overwriting data in the sheet."""
    excel_path, sheet_name, header_row = create_test_excel
    
    # Create a sample Polars DataFrame
    df = generate_test_data({
        "Name": ["Alice", "Bob"],
        "Age": [25, 30],
        "Gender": ["F", "M"]
    }, data_type)

    kwargs = { "overwrite": True }
    if data_type == "list":
        kwargs["columns"] = ["Name", "Age", "Gender"]
        
    template = ExcelTemplate(excel_path)
    template.goto_sheet(sheet_name)
    template.set_header_location((header_row, 1), 'row')
    template.fill_with(df, **kwargs)
    template.save(excel_path)

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheet_name]
    
    # Assert that data is inserted with a header row (named columns)
    assert sheet["A3"].value == "Name"
    assert sheet["B3"].value == "Age"
    assert sheet["C3"].value == "Gender"
    assert sheet["A4"].value == "Alice"
    assert sheet["B4"].value == 25
    assert sheet["C4"].value == "F"
    assert sheet["A5"].value == "Bob"
    assert sheet["B5"].value == 30
    assert sheet["C5"].value == "M"


@pytest.mark.parametrize("data_type", ["polars", "pandas", "dict", "list"])
def test_fill_sheet_with_overwrite_shorter_than_template(create_test_excel, data_type):
    """Test overwriting data in the sheet."""
    excel_path, sheet_name, header_row = create_test_excel
    
    # Create a sample Polars DataFrame
    df = generate_test_data({
        "Name": ["Alice"],
        "Age": [25],
        "Gender": ["F"]
    }, data_type)

    kwargs = { "overwrite": True }
    if data_type == "list":
        kwargs["columns"] = ["Name", "Age", "Gender"]
        
    template = ExcelTemplate(excel_path)
    template.goto_sheet(sheet_name)
    template.set_header_location((header_row, 1), 'row')
    template.fill_with(df, **kwargs)
    template.save(excel_path)

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheet_name]
    
    # Assert that data is inserted with a header row (named columns)
    assert sheet["A3"].value == "Name"
    assert sheet["B3"].value == "Age"
    assert sheet["C3"].value == "Gender"
    assert sheet["A4"].value == "Alice"
    assert sheet["B4"].value == 25
    assert sheet["C4"].value == "F"
    assert sheet["A5"].value == None
    assert sheet["B5"].value == None
    assert sheet["C5"].value == None


@pytest.mark.parametrize("data_type", ["polars", "pandas", "dict", "list"])
def test_fill_sheet_with_overwrite_and_skip_null(create_test_excel, data_type):
    """Test inserting a Polars DataFrame with named columns (`named=True`) and empty values."""
    excel_path, sheet_name, header_row = create_test_excel
    
    # Create a sample Polars DataFrame
    df = generate_test_data({
        "Name": ["Alice", "Bob", "Tom"],
        "Age": [25, None, 30],
        "Gender": ["F", "M", None]
    }, data_type)
    
    kwargs = { "overwrite": True, "skip_null": True }
    if data_type == "list":
        kwargs["columns"] = ["Name", "Age", "Gender"]
        
    template = ExcelTemplate(excel_path)
    template.goto_sheet(sheet_name)
    template.set_header_location((header_row, 1), 'row')
    template.fill_with(df, **kwargs)
    template.save(excel_path)

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert sheet["A3"].value == "Name"
    assert sheet["B3"].value == "Age"
    assert sheet["C3"].value == "Gender"
    assert sheet["A4"].value == "Alice"
    assert sheet["B4"].value == 25
    assert sheet["C4"].value == "F"
    assert sheet["A5"].value == "Bob"
    assert sheet["B5"].value == 26
    assert sheet["C5"].value == "M"
    assert sheet["A6"].value == "Tom"
    assert sheet["B6"].value == 30
    assert sheet["C6"].value == None


def test_fill_sheet_with_empty_column_in_list(create_test_excel):
    """Test behavior when an empty column is in the DataFrame."""
    excel_path, sheet_name, header_row = create_test_excel
    
    # Create a DataFrame with mismatching column names
    df = generate_test_data({
        "Name": ["Alice", "Bob"],
        "Age": [25, 30],
        "Gender": []
    }, 'list')

    kwargs = {}
    kwargs["columns"] = ["Name", "Age", "Gender"]

    template = ExcelTemplate(excel_path)
    template.goto_sheet(sheet_name)
    template.set_header_location((header_row, 1), 'row')
    template.fill_with(df, **kwargs)
    template.save(excel_path)

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert sheet["A3"].value == "Name"
    assert sheet["B3"].value == "Age"
    assert sheet["C3"].value == "Gender"
    assert sheet["A6"].value == "Alice"
    assert sheet["B6"].value == 25
    assert sheet["C6"].value == None
    assert sheet["A7"].value == "Bob"
    assert sheet["B7"].value == 30
    assert sheet["C7"].value == None


@pytest.mark.parametrize("data_type", ["polars", "pandas", "dict", "list"])
def test_fill_sheet_with_multiple_overwrite(create_test_excel, data_type):
    """Test inserting data for specific columns in multiple calls."""
    excel_path, sheet_name, header_row = create_test_excel
    
    # Create a sample Polars DataFrame
    df1 = generate_test_data({
        "Name": ["Alice", "Bob", "Tom"],
        "Age": [25, 30, 35],
    }, data_type)

    kwargs1 = { "overwrite": True }
    if data_type == "list":
        kwargs1["columns"] = ["Name", "Age"]
        
    template = ExcelTemplate(excel_path)
    template.goto_sheet(sheet_name)
    template.set_header_location((header_row, 1), 'row')
    template.fill_with(df1, **kwargs1)

    df2 = generate_test_data({
        "Gender": ["F", "M", "M"]
    }, data_type)

    kwargs2 = { "overwrite": True }
    if data_type == "list":
        kwargs2["columns"] = ["Gender"]
        
    template.set_header_location((header_row, 1), 'row')
    template.fill_with(df2, **kwargs2)
    template.save(excel_path)

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert sheet["A3"].value == "Name"
    assert sheet["B3"].value == "Age"
    assert sheet["C3"].value == "Gender"
    assert sheet["A4"].value == "Alice"
    assert sheet["B4"].value == 25
    assert sheet["C4"].value == "F"
    assert sheet["A5"].value == "Bob"
    assert sheet["B5"].value == 30
    assert sheet["C5"].value == "M"
    assert sheet["A6"].value == "Tom"
    assert sheet["B6"].value == 35
    assert sheet["C6"].value == "M"
