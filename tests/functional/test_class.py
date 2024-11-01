# This project uses Poetry for dependency management.
import pytest
from pathlib import Path
import polars as pl
import pandas as pd
import openpyxl
from ez_excel_mgt import ExcelTemplate


def test_open_and_save_as(create_test_excel):
    """Test opening an existing file and saving as."""
    file_path, _, _ = create_test_excel
    template = ExcelTemplate(file_path)

    assert template is not None

    path = Path(file_path)
    template.save(str(path.with_suffix(".new.xlsx")))

    assert path.with_suffix(".new.xlsx").exists()

    # Clean up
    path.with_suffix(".new.xlsx").unlink()


def test_add_sheet(create_test_excel):
    """Test adding a sheet."""
    file_path, _, _ = create_test_excel
    dest_sheet_name = "Sheet 2"
    template = ExcelTemplate(str(file_path))

    template.add_sheet(dest_sheet_name)
    template.save(file_path)

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[dest_sheet_name]

    assert sheet is not None

def test_list_sheet_names(create_test_excel_with_3_sheets):
    """Test listing sheet names."""
    file_path, sheet_names, _ = create_test_excel_with_3_sheets

    template = ExcelTemplate(file_path)
    
    assert template.sheet_names() == sheet_names

def test_write_cell(create_test_excel):
    """Test writing a cell."""
    file_path, _, _ = create_test_excel
    dest_sheet_name = "Test Sheet"

    template = ExcelTemplate(file_path)
    template.add_sheet(dest_sheet_name)
    template.write_cell(dest_sheet_name, (2, 5), "Hello, World!")
    template.write_cell(dest_sheet_name, "F2", 12345)
    template.save(file_path)

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[dest_sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert sheet["E2"].value == "Hello, World!"
    assert sheet["F2"].value == 12345


def test_set_cell(create_test_excel):
    """Test setting a cell."""
    file_path, _, _ = create_test_excel
    dest_sheet_name = "Test Sheet"

    template = ExcelTemplate(file_path)
    template.add_sheet(dest_sheet_name)
    template.goto_sheet(dest_sheet_name)
    template.goto_cell((5, 2))
    template.set_value("Hello, World!")
    template.goto_cell('B6')
    template.set_value(True)
    template.save(file_path)

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[dest_sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert sheet["B5"].value == "Hello, World!"
    assert sheet["B6"].value == True
