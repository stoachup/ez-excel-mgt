# This project uses Poetry for dependency management.
from pathlib import Path
import openpyxl
from ez_excel_mgt import ExcelTemplate


def test_copy_range_between_files(create_test_excel, create_empty_test_excel):
    """Test copy."""
    source_file_path, source_sheet_name, _header_row = create_test_excel
    dest_file_path, dest_sheet_name, _header_row = create_empty_test_excel

    template = ExcelTemplate(dest_file_path)
    template.goto_sheet(dest_sheet_name, cell=(1, 1))
    template.copy_range_from(source_file_path, source_sheet_name, ((1, 1), (5, 3)), None, None)
    template.save(dest_file_path)

    assert Path(dest_file_path).exists()

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(dest_file_path)
    sheet = workbook[dest_sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert sheet["A1"].value == "First row"
    assert sheet["A2"].value == "Second row"
    assert sheet["B4"].value == 25
    assert sheet["B5"].value == 26

def test_transpose_range_between_files(create_test_excel, create_empty_test_excel):
    """Test copy."""
    source_file_path, source_sheet_name, _header_row = create_test_excel
    dest_file_path, dest_sheet_name, _header_row = create_empty_test_excel

    template = ExcelTemplate(dest_file_path)
    template.goto_sheet(dest_sheet_name, cell=(10, 20))
    template.copy_range_from(source_file_path, source_sheet_name, ((1, 1), (5, 3)), True, None)
    template.save(dest_file_path)

    assert Path(dest_file_path).exists()

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(dest_file_path)
    sheet = workbook[dest_sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert sheet["T10"].value == "First row"
    assert sheet["U10"].value == "Second row"
    assert sheet["W11"].value == 25
    assert sheet["X11"].value == 26


def test_copy_range_between_files_and_coerce(create_test_excel_float, create_empty_test_excel):
    """Test copy."""
    source_file_path, source_sheet_name, _header_row = create_test_excel_float
    dest_file_path, dest_sheet_name, _header_row = create_empty_test_excel

    dest_sheet_name = "New Sheet"

    template = ExcelTemplate(dest_file_path)
    template.add_sheet(dest_sheet_name)
    template.goto_sheet(dest_sheet_name, cell=(1, 1))
    template.copy_range_from(source_file_path, source_sheet_name, ((4, 4), (5, 4)), None, 'int')
    template.save(dest_file_path)

    assert Path(dest_file_path).exists()

    # Load the modified Excel file and verify the contents
    workbook = openpyxl.load_workbook(dest_file_path)
    sheet = workbook[dest_sheet_name]

    # Assert that data is inserted with a header row (named columns)
    assert sheet["A1"].value == 1234
    assert sheet["A2"].value == 9876

